import openpyxl
from openpyxl import load_workbook
import pprint as pp

wb2 = load_workbook('PR_April15.xlsx', read_only=True)


months = ['Jan', 'Feb', 'March', 'April', 'May', 'June',
          'July', 'Aug', 'Sept', 'Oct', 'Nov', 'Dec']

firstMonth = 'Jan15'
currentMonth = 'May15'
lastEntry = 'Over 2021'





def isTitleRow(row):
    return (row[0]. value == 'Comments') and (row[1]. value == 'Contract year')


def getColumnIndex(row):
    colIdx = dict()
    for idx, cell in enumerate(row):
        if cell.value == 'Status':
            colIdx['status'] = idx
        if cell.value == 'Product':
            colIdx['product'] = idx
        if cell.value == 'Project name':
            colIdx['project_name'] = idx
        if cell.value == 'Description':
            if 'description' not in colIdx: # Can have double names
                colIdx['description'] = idx
        if cell.value == 'Maconomy ID':
            colIdx['maconomy_id'] = idx
        if cell.value == 'Consolidate':
            colIdx['consolidate'] = idx
        if cell.value == 'Revenue (TNOK)':
            colIdx['revenue'] = idx
        if cell.value == 'Total hours':
            colIdx['total_hours'] = idx
        if cell.value == 'Expenses (TNOK)':
            colIdx['expenses'] = idx
        if cell.value == 'Total cost (TNOK)':
            colIdx['total_cost'] = idx

        if cell.value == firstMonth:
            if 'first_month' not in colIdx: # Can have double names
                colIdx['first_month'] = idx
        if cell.value == currentMonth:
            if 'current_month' not in colIdx: # Can have double names
                colIdx['current_month'] = idx
        if cell.value == lastEntry:
            if 'last_entry' not in colIdx: # Can have double names
                colIdx['last_entry'] = idx

    return colIdx


print wb2.get_sheet_names()
# sheet1 = wb2['FirstSheet']
sheets = wb2.get_sheet_names()
print sheets
sheet1 = wb2[sheets[0]]
print sheet1['C1'].value

idx = dict()
firstRowFound = False
for row in sheet1.rows:
    if isTitleRow(row):
        firstRowFound = True
        print "First Row !"
        idx = getColumnIndex(row)
        pp.pprint(idx)
        y2015_idx = idx['first_month']
        y2016_idx = y2015_idx + 12
        y2017_idx = y2016_idx + 12
        y2018_idx = y2017_idx + 1
        y2019_idx = y2018_idx + 1
        y2020_idx = y2019_idx + 1
        y2021_idx = y2020_idx + 1
        over2021_idx = y2021_idx + 1

        y2015_length = 12-(idx['current_month']-idx['first_month'])
        y2016_length = 12
        y2017_length = 1
        y2018_length = 1
        y2019_length = 1
        y2020_length = 1
        y2021_length = 1
        over2021_length = 1
    else:
        if firstRowFound:            
            row_i = [(cell.value).encode('utf-8')  for cell in row if cell is not None]

            product = row_i[idx['product']].strip()
            status = row_i[idx['status']].strip()
            description = row_i[idx['description']].strip()

    

    #for cell in row:
    #    pass
        #print(cell.value)
#print sheet1.cell('A','1')
